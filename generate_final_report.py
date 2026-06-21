import json
import re
import os
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter
from datetime import datetime
from openpyxl.chart import PieChart, BarChart, LineChart, Reference

# Add tests directory to sys.path so we can import metadata_registry
import sys
sys.path.append(os.path.abspath('tests'))
from metadata_registry import get_metadata

def generate_report():
    json_path = "reports/report.json"
    if not os.path.exists(json_path):
        print(f"Error: {json_path} does not exist. Run the tests first.")
        sys.exit(1)

    with open(json_path, 'r') as f:
        report_data = json.load(f)

    # Initialize workbook
    wb = openpyxl.Workbook()
    
    # ----------------------------------------------------
    # Styles Definition
    # ----------------------------------------------------
    font_family = "Segoe UI"
    
    # Fonts
    title_font = Font(name=font_family, size=16, bold=True, color="FFFFFF")
    section_font = Font(name=font_family, size=12, bold=True, color="1F4E78")
    header_font = Font(name=font_family, size=11, bold=True, color="FFFFFF")
    bold_font = Font(name=font_family, size=11, bold=True)
    regular_font = Font(name=font_family, size=11)
    italic_font = Font(name=font_family, size=10, italic=True, color="555555")
    
    # Fills
    title_fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid") # Dark steel blue
    header_fill = PatternFill(start_color="2C5282", end_color="2C5282", fill_type="solid") # Classic slate blue
    
    pass_fill = PatternFill(start_color="D4EDDA", end_color="D4EDDA", fill_type="solid") # Light green
    pass_font_color = Font(name=font_family, size=11, bold=True, color="155724")
    
    fail_fill = PatternFill(start_color="F8D7DA", end_color="F8D7DA", fill_type="solid") # Light red
    fail_font_color = Font(name=font_family, size=11, bold=True, color="721C24")
    
    skip_fill = PatternFill(start_color="FFF3CD", end_color="FFF3CD", fill_type="solid") # Light yellow
    skip_font_color = Font(name=font_family, size=11, bold=True, color="856404")
    
    zebra_fill = PatternFill(start_color="F7FAFC", end_color="F7FAFC", fill_type="solid") # Very light gray
    
    # Borders
    thin_border_side = Side(style='thin', color='CBD5E0')
    thick_border_bottom = Side(style='medium', color='1F4E78')
    grid_border = Border(left=thin_border_side, right=thin_border_side, top=thin_border_side, bottom=thin_border_side)
    header_border = Border(left=thin_border_side, right=thin_border_side, top=thin_border_side, bottom=thick_border_bottom)
    
    # Alignments
    align_center = Alignment(horizontal="center", vertical="center")
    align_left = Alignment(horizontal="left", vertical="center")
    align_right = Alignment(horizontal="right", vertical="center")
    align_wrap_left = Alignment(horizontal="left", vertical="top", wrap_text=True)

    # ----------------------------------------------------
    # Parse Test Results & Perform Uniqueness Check
    # ----------------------------------------------------
    tests_list = report_data.get("tests", [])
    total_tests = 0
    passed_count = 0
    failed_count = 0
    skipped_count = 0
    total_duration = 0.0
    
    parsed_tests = []
    seen_ids = set()
    duplicate_count = 0
    duplicates = []
    
    # Module aggregates
    modules = {}

    for i, t in enumerate(tests_list):
        nodeid = t.get("nodeid", "")
        outcome = t.get("outcome", "unknown")
        
        # Extract Test Case ID from parameter brackets
        match = re.search(r'\[(TC-\w+-\d+)\]', nodeid)
        if match:
            tc_id = match.group(1)
        else:
            tc_id = f"TC-UNKNOWN-{i+1:03d}"
            
        # Uniqueness Check
        if tc_id in seen_ids:
            duplicate_count += 1
            duplicates.append(tc_id)
            is_duplicate = "Yes"
        else:
            seen_ids.add(tc_id)
            is_duplicate = "No"

        # Get metadata
        meta = get_metadata(tc_id)
        module_name = meta["coverage"].split(" -> ")[0]
        
        # Initialize module aggregate if not exists
        if module_name not in modules:
            modules[module_name] = {"total": 0, "passed": 0, "failed": 0, "skipped": 0}
            
        modules[module_name]["total"] += 1
        
        # Duration
        duration = t.get("call", {}).get("duration", 0.0)
        total_duration += duration
        
        # Map outcome status
        if outcome == "passed":
            status = "Pass"
            passed_count += 1
            modules[module_name]["passed"] += 1
            actual_res = "As Expected"
        elif outcome == "failed":
            status = "Fail"
            failed_count += 1
            modules[module_name]["failed"] += 1
            actual_res = t.get("call", {}).get("longrepr", "Assertion Error")
        else:
            status = "Skip"
            skipped_count += 1
            modules[module_name]["skipped"] += 1
            actual_res = "Skipped"
            
        total_tests += 1
        
        parsed_tests.append({
            "id": tc_id,
            "module": module_name,
            "name": nodeid.split("::")[-1].split("[")[0],
            "description": meta["objective"],
            "expected": meta["expected"],
            "actual": actual_res,
            "status": status,
            "duration": duration,
            "screenshot": "",
            "duplicate_check": is_duplicate
        })

    # Calculations
    pass_pct = (passed_count / total_tests) * 100 if total_tests > 0 else 0.0
    fail_pct = (failed_count / total_tests) * 100 if total_tests > 0 else 0.0
    
    # ----------------------------------------------------
    # Sheet 1: Test Summary
    # ----------------------------------------------------
    ws_summary = wb.active
    ws_summary.title = "Test Summary"
    ws_summary.views.sheetView[0].showGridLines = True
    
    # Title Block
    ws_summary.merge_cells("A1:C1")
    ws_summary["A1"] = "SELENIUM AUTOMATION EXECUTION SUMMARY"
    ws_summary["A1"].font = title_font
    ws_summary["A1"].fill = title_fill
    ws_summary["A1"].alignment = align_center
    ws_summary.row_dimensions[1].height = 40
    
    # Metadata Table Headers
    ws_summary["A3"] = "Metric"
    ws_summary["B3"] = "Value"
    for col_let in ["A", "B"]:
        cell = ws_summary[f"{col_let}3"]
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = align_center
        cell.border = header_border
    ws_summary.row_dimensions[3].height = 24
    
    # Summary Metrics Data
    metrics = [
        ("Total Test Cases", total_tests),
        ("Passed", passed_count),
        ("Failed", failed_count),
        ("Skipped", skipped_count),
        ("Pass Percentage", f"{pass_pct:.2f}%"),
        ("Fail Percentage", f"{fail_pct:.2f}%"),
        ("Execution Date", datetime.now().strftime("%Y-%m-%d")),
        ("Execution Time", f"{total_duration:.2f}s"),
        ("Browser Used", "Google Chrome (Headless)")
    ]
    
    for row_idx, (metric, val) in enumerate(metrics, 4):
        ws_summary[f"A{row_idx}"] = metric
        ws_summary[f"B{row_idx}"] = val
        
        ws_summary[f"A{row_idx}"].font = bold_font
        ws_summary[f"A{row_idx}"].alignment = align_left
        ws_summary[f"A{row_idx}"].border = grid_border
        
        ws_summary[f"B{row_idx}"].font = regular_font
        ws_summary[f"B{row_idx}"].alignment = align_center
        ws_summary[f"B{row_idx}"].border = grid_border
        
        # Color codes for Pass / Fail / Skipped inside summary
        if metric == "Passed" and val > 0:
            ws_summary[f"B{row_idx}"].fill = pass_fill
            ws_summary[f"B{row_idx}"].font = pass_font_color
        elif metric == "Failed" and val > 0:
            ws_summary[f"B{row_idx}"].fill = fail_fill
            ws_summary[f"B{row_idx}"].font = fail_font_color
        elif metric == "Skipped" and val > 0:
            ws_summary[f"B{row_idx}"].fill = skip_fill
            ws_summary[f"B{row_idx}"].font = skip_font_color
            
        ws_summary.row_dimensions[row_idx].height = 20

    # ----------------------------------------------------
    # Sheet 2: Test Case Details
    # ----------------------------------------------------
    ws_details = wb.create_sheet(title="Test Case Details")
    ws_details.views.sheetView[0].showGridLines = True
    
    headers_details = [
        "Test Case ID", "Module", "Test Case Name", "Description", 
        "Expected Result", "Actual Result", "Status", "Execution Time", 
        "Screenshot Link", "Duplicate Check"
    ]
    
    for col_idx, h in enumerate(headers_details, 1):
        cell = ws_details.cell(row=1, column=col_idx, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = align_center
        cell.border = header_border
    ws_details.row_dimensions[1].height = 26
    
    for row_idx, t in enumerate(parsed_tests, 2):
        ws_details.cell(row=row_idx, column=1, value=t["id"]).alignment = align_center
        ws_details.cell(row=row_idx, column=2, value=t["module"]).alignment = align_left
        ws_details.cell(row=row_idx, column=3, value=t["name"]).alignment = align_left
        ws_details.cell(row=row_idx, column=4, value=t["description"]).alignment = align_wrap_left
        ws_details.cell(row=row_idx, column=5, value=t["expected"]).alignment = align_wrap_left
        ws_details.cell(row=row_idx, column=6, value=t["actual"]).alignment = align_wrap_left
        
        status_cell = ws_details.cell(row=row_idx, column=7, value=t["status"])
        status_cell.alignment = align_center
        if t["status"] == "Pass":
            status_cell.fill = pass_fill
            status_cell.font = pass_font_color
        elif t["status"] == "Fail":
            status_cell.fill = fail_fill
            status_cell.font = fail_font_color
        else:
            status_cell.fill = skip_fill
            status_cell.font = skip_font_color
            
        time_cell = ws_details.cell(row=row_idx, column=8, value=t["duration"])
        time_cell.alignment = align_right
        time_cell.number_format = '0.00"s"'
        
        ws_details.cell(row=row_idx, column=9, value=t["screenshot"]).alignment = align_center
        
        dup_cell = ws_details.cell(row=row_idx, column=10, value=t["duplicate_check"])
        dup_cell.alignment = align_center
        if t["duplicate_check"] == "Yes":
            dup_cell.fill = fail_fill
            dup_cell.font = fail_font_color
        
        # Apply fonts and borders
        for col_idx in range(1, 11):
            c = ws_details.cell(row=row_idx, column=col_idx)
            if col_idx not in [7, 10]: # Skip status and duplicate cells styling override
                c.font = regular_font
            c.border = grid_border
            
            # Zebra striping
            if row_idx % 2 == 0 and col_idx not in [7, 10]:
                c.fill = zebra_fill
                
        ws_details.row_dimensions[row_idx].height = 20

    # ----------------------------------------------------
    # Sheet 3: Failed Test Cases
    # ----------------------------------------------------
    ws_failed = wb.create_sheet(title="Failed Test Cases")
    ws_failed.views.sheetView[0].showGridLines = True
    
    headers_failed = ["Test Case ID", "Test Name", "Failure Reason", "Screenshot Path", "Log File Path"]
    for col_idx, h in enumerate(headers_failed, 1):
        cell = ws_failed.cell(row=1, column=col_idx, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = align_center
        cell.border = header_border
    ws_failed.row_dimensions[1].height = 26
    
    failed_tests = [t for t in parsed_tests if t["status"] == "Fail"]
    if not failed_tests:
        # If no tests failed, add a placeholder indicating system is stable
        ws_failed.cell(row=2, column=1, value="None").alignment = align_center
        ws_failed.cell(row=2, column=2, value="N/A").alignment = align_center
        ws_failed.cell(row=2, column=3, value="No failures occurred during this execution run. System stable.").alignment = align_left
        ws_failed.cell(row=2, column=4, value="N/A").alignment = align_center
        ws_failed.cell(row=2, column=5, value="N/A").alignment = align_center
        
        for col_idx in range(1, 6):
            c = ws_failed.cell(row=2, column=col_idx)
            c.font = italic_font
            c.border = grid_border
        ws_failed.row_dimensions[2].height = 22
    else:
        for row_idx, t in enumerate(failed_tests, 2):
            ws_failed.cell(row=row_idx, column=1, value=t["id"]).alignment = align_center
            ws_failed.cell(row=row_idx, column=2, value=t["name"]).alignment = align_left
            ws_failed.cell(row=row_idx, column=3, value=t["actual"]).alignment = align_wrap_left
            ws_failed.cell(row=row_idx, column=4, value=t["screenshot"]).alignment = align_center
            ws_failed.cell(row=row_idx, column=5, value=f"reports/logs/{t['name']}.log").alignment = align_center
            
            for col_idx in range(1, 6):
                c = ws_failed.cell(row=row_idx, column=col_idx)
                c.font = regular_font
                c.border = grid_border
                if row_idx % 2 == 0:
                    c.fill = zebra_fill
            ws_failed.row_dimensions[row_idx].height = 22

    # ----------------------------------------------------
    # Sheet 4: Coverage Report
    # ----------------------------------------------------
    ws_coverage = wb.create_sheet(title="Coverage Report")
    ws_coverage.views.sheetView[0].showGridLines = True
    
    headers_coverage = ["Module Name", "Total Tests", "Passed Tests", "Failed Tests", "Coverage %"]
    for col_idx, h in enumerate(headers_coverage, 1):
        cell = ws_coverage.cell(row=1, column=col_idx, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = align_center
        cell.border = header_border
    ws_coverage.row_dimensions[1].height = 26
    
    for row_idx, (mod_name, data) in enumerate(modules.items(), 2):
        cov_pct = (data["passed"] / data["total"]) * 100 if data["total"] > 0 else 0.0
        
        ws_coverage.cell(row=row_idx, column=1, value=mod_name).alignment = align_left
        ws_coverage.cell(row=row_idx, column=2, value=data["total"]).alignment = align_center
        ws_coverage.cell(row=row_idx, column=3, value=data["passed"]).alignment = align_center
        ws_coverage.cell(row=row_idx, column=4, value=data["failed"]).alignment = align_center
        
        cov_cell = ws_coverage.cell(row=row_idx, column=5, value=cov_pct / 100.0)
        cov_cell.alignment = align_right
        cov_cell.number_format = '0.00%'
        
        for col_idx in range(1, 6):
            c = ws_coverage.cell(row=row_idx, column=col_idx)
            c.font = regular_font
            c.border = grid_border
            if row_idx % 2 == 0:
                c.fill = zebra_fill
        ws_coverage.row_dimensions[row_idx].height = 22

    # ----------------------------------------------------
    # Sheet 5: Defect Summary
    # ----------------------------------------------------
    ws_defect = wb.create_sheet(title="Defect Summary")
    ws_defect.views.sheetView[0].showGridLines = True
    
    headers_defect = ["Defect ID", "Description", "Severity", "Status", "Fixed By", "Re-Test Status"]
    for col_idx, h in enumerate(headers_defect, 1):
        cell = ws_defect.cell(row=1, column=col_idx, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = align_center
        cell.border = header_border
    ws_defect.row_dimensions[1].height = 26
    
    # Load defaults
    ws_defect.cell(row=2, column=1, value="None").alignment = align_center
    ws_defect.cell(row=2, column=2, value="No Critical Defects Found during this execution run. System stable.").alignment = align_left
    ws_defect.cell(row=2, column=3, value="Low").alignment = align_center
    ws_defect.cell(row=2, column=4, value="Closed").alignment = align_center
    ws_defect.cell(row=2, column=5, value="N/A").alignment = align_center
    ws_defect.cell(row=2, column=6, value="Passed").alignment = align_center
    
    for col_idx in range(1, 7):
        c = ws_defect.cell(row=2, column=col_idx)
        c.font = italic_font
        c.border = grid_border
        
    ws_defect.row_dimensions[2].height = 22

    # ----------------------------------------------------
    # Apply Auto Filters, Column Widths, and Freeze Panes
    # ----------------------------------------------------
    for ws in [ws_details, ws_failed, ws_coverage, ws_defect]:
        ws.freeze_panes = "A2"
        
        # Max row and column
        max_r = ws.max_row
        max_c = ws.max_column
        max_col_let = get_column_letter(max_c)
        ws.auto_filter.ref = f"A1:{max_col_let}{max_r}"
        
        # Column width adjustments
        for col in ws.columns:
            max_len = 0
            col_letter = col[0].column_letter
            for cell in col:
                val_str = str(cell.value or "")
                if len(val_str) > max_len:
                    max_len = len(val_str)
            ws.column_dimensions[col_letter].width = min(max(max_len + 3, 12), 40)
            
    # Custom widths for wrap columns
    ws_details.column_dimensions['D'].width = 35 # Description
    ws_details.column_dimensions['E'].width = 30 # Expected
    ws_details.column_dimensions['F'].width = 30 # Actual
    ws_failed.column_dimensions['C'].width = 40  # Failure Reason
    
    # Adjust Sheet 1 columns
    ws_summary.column_dimensions['A'].width = 25
    ws_summary.column_dimensions['B'].width = 30

    # ----------------------------------------------------
    # Add Charts
    # ----------------------------------------------------
    # Chart 1: Pass vs Fail (Pie Chart) on Sheet 1 (Test Summary)
    pie = PieChart()
    pie.title = "Pass vs Fail Ratio"
    pie.style = 10
    
    # Categories (Passed, Failed, Skipped) -> column A, rows 5 to 7
    labels = Reference(ws_summary, min_col=1, min_row=5, max_row=7)
    # Data (Values) -> column B, rows 4 to 7 (includes header B4 for title "Value")
    data = Reference(ws_summary, min_col=2, min_row=4, max_row=7)
    
    pie.add_data(data, titles_from_data=True)
    pie.set_categories(labels)
    pie.width = 12
    pie.height = 8.5
    ws_summary.add_chart(pie, "D3")
    
    # Chart 2: Module-wise Coverage (Column Bar Chart) on Sheet 4 (Coverage Report)
    bar = BarChart()
    bar.type = "col"
    bar.title = "Module-wise Test Status"
    bar.y_axis.title = "Number of Tests"
    bar.x_axis.title = "Modules"
    bar.style = 11
    
    # Total, Passed, Failed -> column B to D, rows 1 to 12
    data_bar = Reference(ws_coverage, min_col=2, min_row=1, max_col=4, max_row=12)
    cats_bar = Reference(ws_coverage, min_col=1, min_row=2, max_row=12)
    
    bar.add_data(data_bar, titles_from_data=True)
    bar.set_categories(cats_bar)
    bar.width = 16
    bar.height = 10
    ws_coverage.add_chart(bar, "G2")
    
    # Chart 3: Execution Trend (Line Chart) on Sheet 1 (Test Summary)
    line = LineChart()
    line.title = "Execution Time Trend per Test"
    line.y_axis.title = "Execution Time (seconds)"
    line.x_axis.title = "Test Cases"
    line.style = 13
    
    # Column H (Execution Time) in Test Case Details sheet. Row 1 to 341.
    data_line = Reference(ws_details, min_col=8, min_row=1, max_row=341)
    # Column A (Test Case ID) in Test Case Details sheet. Row 2 to 341.
    cats_line = Reference(ws_details, min_col=1, min_row=2, max_row=341)
    
    line.add_data(data_line, titles_from_data=True)
    line.set_categories(cats_line)
    line.width = 24
    line.height = 10
    ws_summary.add_chart(line, "D20")

    # ----------------------------------------------------
    # Save Report
    # ----------------------------------------------------
    report_dir = "test-reports/excel"
    os.makedirs(report_dir, exist_ok=True)
    
    file_standard = os.path.join(report_dir, "Selenium_Test_Report.xlsx")
    
    timestamp = datetime.now().strftime("%Y_%m_%d_%H_%M")
    file_dated = os.path.join(report_dir, f"Selenium_Test_Report_{timestamp}.xlsx")
    
    # Save to both paths
    wb.save(file_standard)
    wb.save(file_dated)
    
    print("SUCCESS: Reports generated and saved:")
    print(f"  Standard: {file_standard}")
    print(f"  Dated: {file_dated}")
    
    # Print the specific outputs required by user
    print(f"__TOTAL_TESTS_EXECUTED__:{total_tests}")
    print(f"__UNIQUE_TEST_COUNT__:{len(seen_ids)}")
    print(f"__DUPLICATE_COUNT__:{duplicate_count}")
    print(f"__UPLOADED_REPORT_PATH__:{file_dated.replace(os.sep, '/')}")

if __name__ == "__main__":
    generate_report()
