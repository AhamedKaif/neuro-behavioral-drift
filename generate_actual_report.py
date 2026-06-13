import json
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
import ast
import os

def get_docstrings(directory):
    """Scan directory for python files and extract docstrings for functions."""
    docstrings = {}
    for root, _, files in os.walk(directory):
        for file in files:
            if file.endswith('.py') and file.startswith('test_'):
                filepath = os.path.join(root, file)
                try:
                    with open(filepath, 'r', encoding='utf-8') as f:
                        source = f.read()
                    module = ast.parse(source)
                    for node in ast.walk(module):
                        if isinstance(node, ast.FunctionDef) and node.name.startswith('test_'):
                            docstring = ast.get_docstring(node)
                            # key: filepath/node.name
                            rel_path = os.path.relpath(filepath, directory).replace('\\', '/')
                            docstrings[f"tests/{rel_path}::{node.name}"] = docstring or ""
                            docstrings[node.name] = docstring or ""
                except Exception as e:
                    print(f"Error parsing {filepath}: {e}")
    return docstrings

def style_header(ws):
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill("solid", fgColor="4F81BD")
    alignment = Alignment(horizontal="center", vertical="center")
    thin = Side(border_style="thin", color="000000")
    border = Border(top=thin, left=thin, right=thin, bottom=thin)
    
    for cell in ws[1]:
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = alignment
        cell.border = border

def style_data(ws):
    pass_fill = PatternFill("solid", fgColor="C6EFCE")
    fail_fill = PatternFill("solid", fgColor="FFC7CE")
    pass_font = Font(color="006100", bold=True)
    fail_font = Font(color="9C0006", bold=True)
    alignment = Alignment(wrap_text=True, vertical="center")
    thin = Side(border_style="thin", color="000000")
    border = Border(top=thin, left=thin, right=thin, bottom=thin)
    
    for row in ws.iter_rows(min_row=2):
        for cell in row:
            cell.alignment = alignment
            cell.border = border
            if cell.column == 7: # Status column
                if cell.value == "PASS":
                    cell.fill = pass_fill
                    cell.font = pass_font
                elif cell.value == "FAIL":
                    cell.fill = fail_fill
                    cell.font = fail_font

def set_col_widths(ws):
    for col in ws.columns:
        max_length = 0
        column = col[0].column_letter
        for cell in col:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        ws.column_dimensions[column].width = min(max_length + 2, 50)

def generate_report():
    report_file = 'report.json'
    if not os.path.exists(report_file):
        print(f"Error: {report_file} not found. Please run pytest with --json-report first.")
        return

    with open(report_file, 'r', encoding='utf-8') as f:
        report_data = json.load(f)

    docstrings = get_docstrings('tests')
    
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Executed Selenium Tests"

    # Insert Summary later, let's leave rows 1-6 for summary
    ws.append([]) # row 1
    ws.append([]) # row 2
    ws.append([]) # row 3
    ws.append([]) # row 4
    ws.append([]) # row 5
    ws.append([]) # row 6
    
    headers = ["S.No", "Test Case ID", "Test File Name", "Test Method Name", "Module", "Description", "Execution Status", "Execution Time", "Error Message"]
    ws.append(headers) # row 7

    tests = report_data.get('tests', [])
    
    passed_count = 0
    failed_count = 0
    
    for idx, test in enumerate(tests, 1):
        nodeid = test.get('nodeid', '')
        # Node id format: tests/test_login.py::test_valid_login
        if '::' in nodeid:
            file_path, method_name = nodeid.split('::', 1)
        else:
            file_path = nodeid
            method_name = nodeid
            
        file_name = os.path.basename(file_path)
        
        # Determine Module
        module_name = file_name.replace('test_', '').replace('.py', '').capitalize()
        
        # Description
        desc = docstrings.get(method_name, "")
        if not desc:
            desc = method_name.replace('test_', '').replace('_', ' ').capitalize()
            
        # Status
        outcome = test.get('outcome', '')
        if outcome == 'passed':
            status = 'PASS'
            passed_count += 1
        elif outcome == 'failed':
            status = 'FAIL'
            failed_count += 1
        else:
            status = outcome.upper()
            
        # Time
        call = test.get('call', {})
        duration = call.get('duration', 0.0)
        time_str = f"{duration:.2f}s"
        
        # Error Message
        error_msg = ""
        if status == 'FAIL':
            error_msg = call.get('crash', {}).get('message', '')
            if not error_msg:
                error_msg = call.get('longrepr', '')[:200] # Trim long errors
                
        tc_id = f"TC_{idx:03d}"
        
        row_data = [
            idx,
            tc_id,
            file_name,
            method_name,
            module_name,
            desc,
            status,
            time_str,
            error_msg
        ]
        ws.append(row_data)

    total_tests = len(tests)
    pass_pct = f"{(passed_count/total_tests)*100:.2f}%" if total_tests > 0 else "0%"
    
    # Write summary at the top
    ws['A1'] = "Testing Summary"
    ws['A1'].font = Font(bold=True, size=14)
    
    ws['A2'] = "Total Tests Executed"
    ws['B2'] = total_tests
    
    ws['A3'] = "Total Passed"
    ws['B3'] = passed_count
    
    ws['A4'] = "Total Failed"
    ws['B4'] = failed_count
    
    ws['A5'] = "Pass Percentage"
    ws['B5'] = pass_pct
    
    for r in range(2, 6):
        ws[f'A{r}'].font = Font(bold=True)

    # Style header at row 7
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill("solid", fgColor="4F81BD")
    alignment = Alignment(horizontal="center", vertical="center")
    thin = Side(border_style="thin", color="000000")
    border = Border(top=thin, left=thin, right=thin, bottom=thin)
    
    for cell in ws[7]:
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = alignment
        cell.border = border
        
    # Style data
    pass_fill = PatternFill("solid", fgColor="C6EFCE")
    fail_fill = PatternFill("solid", fgColor="FFC7CE")
    pass_font = Font(color="006100", bold=True)
    fail_font = Font(color="9C0006", bold=True)
    
    for row in ws.iter_rows(min_row=8):
        for cell in row:
            cell.alignment = Alignment(wrap_text=True, vertical="center")
            cell.border = border
            if cell.column == 7: # Status column
                if cell.value == "PASS":
                    cell.fill = pass_fill
                    cell.font = pass_font
                elif cell.value == "FAIL":
                    cell.fill = fail_fill
                    cell.font = fail_font

    set_col_widths(ws)

    filename = "Executed Selenium Tests.xlsx"
    file_path = os.path.abspath(filename)
    wb.save(file_path)
    print(f"Generated successfully: {file_path}")

if __name__ == "__main__":
    generate_report()
